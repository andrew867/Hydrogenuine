"""Pack 14: XLSX export builder — create, add sheet/data, finalize to tenant exports. Returns file_id for download."""

from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any, Dict, List, Union

from hg_core.docs.paths import get_exports_root

_xlsx_buffers: Dict[str, Any] = {}


def xlsx_create(title: str, tenant_id: str) -> str:
    """Create a new XLSX workbook buffer. Returns doc_id for add_sheet / add_data / finalize."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required; pip install openpyxl")
    doc_id = str(uuid.uuid4())
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws:
        ws.title = (title or "Sheet1")[:31]
    _xlsx_buffers[doc_id] = {"wb": wb, "tenant_id": tenant_id}
    return doc_id


def xlsx_add_sheet(doc_id: str, sheet_name: str, rows: List[List[Any]]) -> None:
    """Add a sheet (or create and append rows)."""
    buf = _xlsx_buffers.get(doc_id)
    if not buf:
        raise ValueError("doc_id not found")
    wb = buf["wb"]
    ws = wb.create_sheet(title=(sheet_name or "Sheet")[:31])
    for r in rows:
        ws.append(r if isinstance(r, (list, tuple)) else [r])


def xlsx_add_data(doc_id: str, rows: List[List[Any]], sheet_name: str = "Data") -> None:
    """Append rows to the active (first) sheet or create one."""
    buf = _xlsx_buffers.get(doc_id)
    if not buf:
        raise ValueError("doc_id not found")
    wb = buf["wb"]
    ws = wb.active
    if not ws:
        ws = wb.create_sheet(title=sheet_name[:31])
    for r in rows:
        ws.append(r if isinstance(r, (list, tuple)) else [r])


def xlsx_finalize(doc_id: str, filename: str) -> str:
    """Write XLSX to tenant exports dir. Returns file_id for GET /v1/files/{file_id}/download."""
    buf = _xlsx_buffers.pop(doc_id, None)
    if not buf:
        raise ValueError("doc_id not found")
    tenant_id = buf["tenant_id"]
    export_root = get_exports_root(tenant_id)
    export_root.mkdir(parents=True, exist_ok=True)
    file_id = str(uuid.uuid4())
    safe_name = (filename or "export").strip().replace("..", "") or "export"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    path = export_root / f"{file_id}.xlsx"
    buf["wb"].save(str(path))
    return file_id
